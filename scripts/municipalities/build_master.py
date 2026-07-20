"""総務省「全国地方公共団体コード」(都道府県コード及び市区町村コード)を読み込み、
市町村コードマスタを生成する(SPEC.md 5章、docs/国保料データ全国展開_設計方針_2026-07-19.md フェーズ1)。

一次ソース: data/raw/municipalities/000925835.xlsx(総務省公表、令和6年1月1日時点)
  シート「R6.1.1現在の団体」: 基礎自治体(市・町・村・特別区)の一覧。政令指定都市は
    区に分割されず1件の市として掲載される。ここから src/data/municipalities/master.json
    (1,747件)を生成する。
  シート「R6.1.1政令指定都市」: 政令指定都市とその区の内訳。区の行(市区町村名が「区」で
    終わる行)だけを src/data/municipalities/designated-city-wards.json(参考データ、
    上記1,747件には含めない)として生成する。国保料は市単位で算定されるため、区は
    parentCode で親市の5桁コードを参照する。

団体コードは6桁(JIS X 0402の5桁コード+検査数字1桁)で記載されているため、
先頭5桁を市区町村コードとして採用する(既存の src/data/rates/2026/national-health-insurance/
{municipalityCode}.json 等のファイル命名規則と一致させるため)。

使い方:
    python scripts/municipalities/build_master.py
"""
import json
import unicodedata
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
SOURCE_XLSX = ROOT / "data" / "raw" / "municipalities" / "000925835.xlsx"
MASTER_OUT = ROOT / "src" / "data" / "municipalities" / "master.json"
WARDS_OUT = ROOT / "src" / "data" / "municipalities" / "designated-city-wards.json"

MUNICIPALITIES_SHEET = "R6.1.1現在の団体"
DESIGNATED_CITY_SHEET = "R6.1.1政令指定都市"


def to_code5(raw_code) -> str:
    code6 = str(raw_code).zfill(6)
    return code6[:5]


def to_kana(value: str) -> str:
    """半角カナを全角カナに正規化する(入力補完・検索用の読み仮名として使うため)。"""
    return unicodedata.normalize("NFKC", value)


def type_of(name: str) -> str:
    if name.endswith("区"):
        return "特別区"
    if name.endswith("市"):
        return "市"
    if name.endswith("町"):
        return "町"
    if name.endswith("村"):
        return "村"
    raise ValueError(f"未知の市区町村種別: {name}")


def build_master(wb) -> list[dict]:
    ws = wb[MUNICIPALITIES_SHEET]
    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_code, pref, muni, pref_kana, muni_kana = row[0], row[1], row[2], row[3], row[4]
        if raw_code is None or muni is None:
            continue  # 都道府県自体の行(市区町村名が空)は除外する
        code = to_code5(raw_code)
        entries.append(
            {
                "code": code,
                "prefecture": pref,
                "prefectureCode": code[:2],
                "municipality": muni,
                "municipalityKana": to_kana(muni_kana),
                "type": type_of(muni),
            }
        )
    entries.sort(key=lambda e: e["code"])
    return entries


def build_wards(wb) -> list[dict]:
    ws = wb[DESIGNATED_CITY_SHEET]
    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_code, pref, muni, pref_kana, muni_kana = row[0], row[1], row[2], row[3], row[4]
        if raw_code is None or muni is None or not muni.endswith("区"):
            continue  # 区の行だけを対象にし、政令指定都市自体の行(親市)は除外する
        code = to_code5(raw_code)
        entries.append(
            {
                "code": code,
                "prefecture": pref,
                "municipality": muni,
                "municipalityKana": to_kana(muni_kana),
                "type": "政令指定都市の区",
            }
        )
    return entries


def attach_parent_codes(wards: list[dict], master: list[dict]) -> None:
    """区の市区町村名(例: 札幌市中央区)から親市名(例: 札幌市)を切り出し、
    master.json 上のコードに解決する。"""
    city_code_by_name = {(e["prefecture"], e["municipality"]): e["code"] for e in master}
    for ward in wards:
        pref = ward["prefecture"]
        matched_city = None
        for (p, city_name) in city_code_by_name:
            if p == pref and ward["municipality"].startswith(city_name) and city_name != ward["municipality"]:
                if matched_city is None or len(city_name) > len(matched_city):
                    matched_city = city_name
        if matched_city is None:
            raise ValueError(f"親市が特定できません: {ward}")
        ward["parentCode"] = city_code_by_name[(pref, matched_city)]


def main():
    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True)
    master = build_master(wb)
    wards = build_wards(wb)
    attach_parent_codes(wards, master)

    MASTER_OUT.parent.mkdir(parents=True, exist_ok=True)
    MASTER_OUT.write_text(
        json.dumps(
            {
                "description": "総務省「全国地方公共団体コード」(都道府県コード及び市区町村コード、令和6年1月1日時点、data/raw/municipalities/000925835.xlsx)から生成した基礎自治体マスタ。政令指定都市の区はここに含めず designated-city-wards.json に分離する。収集進捗・検証状況は src/data/municipalities/index.json 側で管理する。",
                "source": "data/raw/municipalities/source.json",
                "count": len(master),
                "municipalities": master,
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )
    print(f"[build_master] master.json: {len(master)}件")

    WARDS_OUT.write_text(
        json.dumps(
            {
                "description": "政令指定都市の区の一覧(総務省公表データより)。国保料は市単位で算定されるため、master.jsonの1,747件には含めない。parentCodeで親市のcodeを参照する。",
                "count": len(wards),
                "wards": wards,
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )
    print(f"[build_master] designated-city-wards.json: {len(wards)}件")


if __name__ == "__main__":
    main()
